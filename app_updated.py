from flask_sqlalchemy import SQLAlchemy
from flask_login import UserMixin, LoginManager, login_user, logout_user, current_user, login_required
from flask import Flask, render_template, redirect, url_for, request, flash, send_file, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask_migrate import Migrate
from datetime import datetime
import os
import json
import io
import xlsxwriter
import pandas as pd
import base64
import logging
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

db = SQLAlchemy()

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    sheets = db.relationship('Sheet', backref='user', lazy=True)
    downloads = db.relationship('DownloadHistory', backref='user', lazy=True)

class Sheet(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sheet_name = db.Column(db.String(100), nullable=False)
    data = db.Column(db.JSON, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class DownloadHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    sheet_data = db.Column(db.JSON)
    downloaded_at = db.Column(db.DateTime, default=datetime.utcnow)

# Load environment variables
load_dotenv()

# Initialize app
app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key'

# Database configuration for both local development and Render deployment
import os

# for local dev
db_file = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'app.db')

# for Render deploy (save db inside /tmp/)
if os.environ.get('RENDER'):
    db_file = '/tmp/app.db'

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + db_file
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.logger.setLevel(logging.DEBUG)

# Configuration for uploads
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['ALLOWED_EXTENSIONS'] = ALLOWED_EXTENSIONS  # Use the variable defined above

# Initialize login manager - Move this section before using any login functionality
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Initialize database
db.init_app(app)
migrate = Migrate(app, db)

# Ensure database tables exist
with app.app_context():
    db.create_all()

# Add this after your app initialization
@app.context_processor
def inject_templates():
    return {
        'built_in_templates': built_in_templates,
        'custom_templates': custom_templates
    }

@app.template_filter('char')
def to_char(value):
    """Convert a number to a character (0->A, 1->B, etc.)"""
    return chr(value)

@app.template_filter('column_letter')
def column_letter(num):
    """Convert a number to Excel-style column letter (0=A, 1=B, etc.)"""
    letters = ''
    while num >= 0:
        letters = chr(65 + (num % 26)) + letters
        num = num // 26 - 1
        if num < 0:
            break
    return letters

@app.template_filter('enumerate')
def _enumerate(sequence):
    """Add enumerate function to Jinja templates"""
    return enumerate(sequence)

# Global data stores
sheets_data = []
custom_templates = []
sheet_order = []

# Built-in templates
built_in_templates = [
    {
        "id": "front_page",
        "name": "Front Page",
        "default_sheet_name": "Front Page",
        "headers": ["Title", "Subtitle", "Date", "Author"],
        "sample_data": [
            ["Course Workbook", "Comprehensive Guide", "2023-09-01", "John Doe"],
            ["", "Module 1", "", ""],
            ["", "Module 2", "", ""]
        ],
        "formatting": {
            "title_font_size": 20,
            "title_bold": True,
            "header_bg_color": "#4F81BD",
            "cell_font_size": 12,
            "cell_alignment": "center",
            "cell_border_color": "#000000",
            "text_wrap": True,
            "merge_cells": "A1:D1,A2:D2,A3:D3"
        }
    },
    {
        "id": "summary",
        "name": "Summary Sheet",
        "default_sheet_name": "Summary",
        "headers": ["Item", "Value", "Formula", "Notes"],
        "sample_data": [
            ["Total Students", "=COUNT(Data!B2:B100)", "", "Automatic count"],
            ["Average Score", "=AVERAGE(Data!C2:C100)", "", "Calculated field"]
        ],
        "formatting": {
            "header_bg_color": "#9BBB59",
            "title_font_size": 14,
            "cell_alignment": "left",
            "number_format": "#,##0.00",
            "validation": {"Notes": {"type": "list", "options": ["Important", "Review", "Done"]}}
        }
    },
    {
        'id': 'course_file',
        'name': 'Course File Template',
        'headers': ['Subject', 'Code', 'Class', 'Semester', 'Teacher'],
        'sample_data': [
            ['Distributed Computing', 'CSC801', 'BE', 'VIII', 'Dr. K. T. Patil']
        ]
    },
    {
        'id': 'attendance',
        'name': 'Attendance Sheet',
        'headers': ['Roll No', 'Name', 'Date 1', 'Date 2', 'Date 3', 'Percentage'],
        'sample_data': [
            ['1', 'Student 1', 'P', 'P', 'A', '66.67%'],
            ['2', 'Student 2', 'P', 'P', 'P', '100%']
        ]
    },
    {
        'id': 'marks',
        'name': 'Marks Sheet',
        'headers': ['Roll No', 'Name', 'Quiz 1', 'Quiz 2', 'Final', 'Total'],
        'sample_data': [
            ['1', 'Student 1', '8', '9', '85', '102'],
            ['2', 'Student 2', '9', '10', '88', '107']
        ]
    }
]

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def get_column_letter_index(column_letter):
    """Convert Excel column letter to index"""
    result = 0
    for char in column_letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1

def get_cell_format(workbook, formatting, row, col):
    """Create cell format based on stored formatting"""
    cell_key = f"{row + 1},{col + 1}"
    cell_format = workbook.add_format()
    
    if cell_key in formatting.get('formatted_cells', {}):
        fmt = formatting['formatted_cells'][cell_key]
        
        # Apply font formatting
        if 'font' in fmt:
            font = fmt['font']
            if font:
                cell_format.set_font_name(font.get('name', 'Calibri'))
                cell_format.set_font_size(float(font.get('size', 11)))
                cell_format.set_bold(font.get('bold', False))
                cell_format.set_italic(font.get('italic', False))
                
                # Handle font color
                if 'color' in font and font['color']:
                    color = font['color'].get('rgb', '')
                    if color:
                        cell_format.set_font_color(color)
        
        # Apply fill
        if 'fill' in fmt:
            fill = fmt['fill']
            if fill and 'fgColor' in fill:
                color = fill['fgColor'].get('rgb', '')
                if color:
                    cell_format.set_pattern(1)
                    cell_format.set_bg_color(color)
        
        # Apply alignment
        if 'alignment' in fmt:
            alignment = fmt['alignment']
            if alignment:
                cell_format.set_align(alignment.get('horizontal', 'left'))
                cell_format.set_align(alignment.get('vertical', 'top'))
        
        # Apply border
        if 'border' in fmt:
            border = fmt['border']
            if border:
                for side in ['top', 'right', 'bottom', 'left']:
                    if side in border:
                        style = border[side].get('style', 'thin')
                        color = border[side].get('color', {}).get('rgb', '#000000')
                        cell_format.set_border(get_border_style(style))
                        cell_format.set_border_color(color)
        
        # Apply number format
        if 'number_format' in fmt:
            cell_format.set_num_format(fmt['number_format'])
            
    return cell_format

def get_border_style(style_name):
    """Convert border style names to xlsxwriter values"""
    style_map = {
        'thin': 1,
        'medium': 2,
        'thick': 3,
        'dotted': 4,
        'dashed': 5,
        'double': 6,
    }
    return style_map.get(style_name, 1)

def serialize_style_object(obj):
    """Convert style objects to serializable dictionaries"""
    if obj is None:
        return None
    
    if isinstance(obj, bytes):
        return serialize_binary_data(obj)
    
    if hasattr(obj, '__dict__'):
        # Convert object attributes to dictionary
        result = {}
        for key, value in obj.__dict__.items():
            if not key.startswith('_'):  # Skip private attributes
                if isinstance(value, bytes):
                    result[key] = serialize_binary_data(value)
                elif hasattr(value, '__dict__'):
                    result[key] = serialize_style_object(value)
                else:
                    result[key] = str(value) if not isinstance(value, (bool, int, float)) else value
        return result
    return str(obj)

def serialize_binary_data(data):
    """Convert binary data to base64 string"""
    if isinstance(data, bytes):
        return base64.b64encode(data).decode('utf-8')
    return data

def serialize_color(color_obj):
    """Convert openpyxl color objects to hex string"""
    if color_obj is None:
        return None
    if hasattr(color_obj, 'rgb'):
        if isinstance(color_obj.rgb, str):
            return color_obj.rgb
        elif isinstance(color_obj.rgb, bytes):
            return color_obj.rgb.hex()
    return None

def serialize_cell_format(cell):
    """Convert cell formatting to serializable dictionary"""
    return {
        'font': {
            'name': cell.font.name or 'Arial',
            'size': cell.font.size or 11,
            'bold': cell.font.bold,
            'italic': cell.font.italic,
            'color': serialize_color(cell.font.color)
        },
        'alignment': {
            'horizontal': cell.alignment.horizontal or 'left',
            'vertical': cell.alignment.vertical or 'center',
            'wrap_text': cell.alignment.wrap_text
        },
        'border': {
            'style': cell.border.outline_level if cell.border else 1,
            'double': any(getattr(cell.border, side).style == 'double' 
                         for side in ['top', 'right', 'bottom', 'left'] 
                         if getattr(cell.border, side))
        },
        'fill': {
            'type': cell.fill.fill_type if cell.fill else None,
            'color': serialize_color(cell.fill.fgColor) if cell.fill else None
        }
    }

@app.route('/')
def root():
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    return redirect(url_for('index'))

@app.route('/index')
@login_required
def index():
    return render_template(
        'index.html',
        sheets=current_user.sheets,
        built_in_templates=built_in_templates,
        custom_templates=custom_templates
    )

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
        
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        remember = request.form.get('remember') == 'on'
        
        user = User.query.filter_by(email=email).first()
        
        if user and check_password_hash(user.password, password):
            login_user(user, remember=remember)  # Set remember=True to maintain login
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('index'))
        
        flash('Invalid email or password', 'danger')
    
    return render_template('login.html', 
                        built_in_templates=built_in_templates, 
                        custom_templates=custom_templates)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
        
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        
        # Validate input
        if not all([name, email, password]):
            flash('All fields are required')
            return redirect(url_for('register'))
        
        # Check if user exists
        if User.query.filter_by(email=email).first():
            flash('Email already registered')
            return redirect(url_for('register'))
        
        try:
            # Create new user
            hashed_password = generate_password_hash(password)
            user = User(name=name, email=email, password=hashed_password)
            db.session.add(user)
            db.session.commit()
            
            # Log in the new user
            login_user(user)
            flash('Registration successful!')
            return redirect(url_for('index'))
            
        except Exception as e:
            app.logger.error(f"Registration error: {str(e)}")
            db.session.rollback()
            flash('Registration failed. Please try again.')
            return redirect(url_for('register'))
    
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/add_sheet', methods=['POST'])
@login_required
def add_sheet():
    try:
        sheet_name = request.form.get('sheet_name', '').strip()
        data_raw = request.form.get('data', '').strip()
        
        if not sheet_name:
            flash('Sheet name is required!', 'error')
            return redirect(url_for('index'))
            
        if Sheet.query.filter_by(user_id=current_user.id, sheet_name=sheet_name).first():
            flash(f'Sheet "{sheet_name}" already exists!', 'error')
            return redirect(url_for('index'))

        # Parse data and ensure it's in the correct format
        data = []
        for row in data_raw.split(';'):
            row = row.strip()
            if row:
                row_data = []
                for cell in row.split(','):
                    cell = cell.strip()
                    # Preserve formulas
                    if cell.startswith('='):
                        row_data.append(cell)
                    else:
                        # Try to convert to number if possible
                        try:
                            if '.' in cell:
                                value = float(cell)
                            else:
                                value = int(cell)
                            row_data.append(value)
                        except ValueError:
                            row_data.append(cell)
                data.append(row_data)

        if not data:
            flash('No data provided!', 'error')
            return redirect(url_for('index'))

        # Store data as JSON string
        new_sheet = Sheet(
            sheet_name=sheet_name,
            data=json.dumps(data),
            user_id=current_user.id
        )
        db.session.add(new_sheet)
        db.session.commit()

        flash(f'Sheet "{sheet_name}" added!', 'success')
        return redirect(url_for('index'))

    except Exception as e:
        app.logger.error(f"Error adding sheet: {str(e)}")
        flash(f'Error adding sheet: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/remove_sheet/<sheet_name>', methods=['POST'])
@login_required
def remove_sheet(sheet_name):
    sheet = Sheet.query.filter_by(user_id=current_user.id, sheet_name=sheet_name).first()
    
    if sheet:
        db.session.delete(sheet)
        db.session.commit()
        flash(f'Sheet "{sheet_name}" removed successfully!', 'info')
    else:
        flash(f'No sheet named "{sheet_name}" found.', 'error')
    return redirect(url_for('index'))

@app.route('/reorder_sheets', methods=['POST'])
@login_required
def reorder_sheets():
    global sheet_order
    new_order = request.form.getlist('order[]')
    if new_order:
        sheet_order = new_order
        flash('Sheet order updated successfully!', 'success')
        return redirect(url_for('index'))
    flash('Failed to update sheet order.', 'error')
    return redirect(url_for('index'))

@app.route('/add_custom_template', methods=['POST'])
@login_required
def add_custom_template():
    try:
        template_id = request.form.get('template_id', '').strip()
        name = request.form.get('name', '').strip()
        default_sheet_name = request.form.get('default_sheet_name', '').strip()
        headers_raw = request.form.get('headers', '').strip()
        sample_data_raw = request.form.get('sample_data', '').strip()
        
        if not all([template_id, name, default_sheet_name, headers_raw]):
            flash("Missing required fields for custom template.", "error")
            return redirect(url_for('index'))
        
        formatting = {
            "header_bg_color": request.form.get('header_bg_color', "#FFFFFF"),
            "title_font_size": int(request.form.get('title_font_size', 12)),
            "title_bold": request.form.get('title_bold', 'off') == 'on',
            "cell_font_size": int(request.form.get('cell_font_size', 11)),
            "cell_alignment": request.form.get('cell_alignment', 'left'),
            "cell_border_color": request.form.get('cell_border_color', "#D3D3D3"),
            "text_wrap": request.form.get('text_wrap', 'off') == 'on',
            "number_format": request.form.get('number_format', 'General'),
            "merge_cells": request.form.get('merge_cells', ''),
            "freeze_panes": request.form.get('freeze_panes', '')
        }
        
        headers = [h.strip() for h in headers_raw.split(',') if h.strip()]
        sample_data = []
        if sample_data_raw:
            for row in sample_data_raw.split('\n'):
                if row.strip():
                    values = [v.strip() for v in row.split('\t')]
                    sample_data.append(values)
        
        new_template = {
            "id": template_id,
            "name": name,
            "default_sheet_name": default_sheet_name,
            "headers": headers,
            "sample_data": sample_data,
            "formatting": formatting
        }
        
        custom_templates.append(new_template)
        flash(f'Custom template "{name}" added successfully!', 'success')
        
    except Exception as e:
        app.logger.error(f"Error adding template: {str(e)}")
        flash('Error adding template. Please check your inputs.', 'error')
    
    return redirect(url_for('index'))

@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    try:
        if 'file' not in request.files:
            app.logger.error("No file part in request")
            flash('No file submitted', 'error')
            return redirect(url_for('index'))
            
        file = request.files['file']
        
        if file.filename == '':
            app.logger.error("Empty filename submitted")
            flash('No file selected', 'error')
            return redirect(url_for('index'))

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            # Log file details
            app.logger.info(f"Processing file: {filename}")
            app.logger.info(f"File size: {os.fstat(file.fileno()).st_size} bytes")
            
            file.save(filepath)
            
            # Verify file was saved
            if not os.path.exists(filepath):
                app.logger.error(f"File not saved to {filepath}")
                flash('Error saving file', 'error')
                return redirect(url_for('index'))
                
            # Read and validate Excel data
            df = pd.read_excel(filepath, sheet_name=None)
            
            if not df:
                app.logger.error("No data found in Excel file")
                flash('Excel file appears to be empty', 'error')
                return redirect(url_for('index'))
                
            sheets_data = []
            sheet_order = []
            
            for sheet_name, sheet_df in df.items():
                if sheet_df.empty:
                    app.logger.warning(f"Sheet '{sheet_name}' is empty")
                    continue
                    
                sheet_data = {
                    'sheet_name': sheet_name,
                    'data': [sheet_df.columns.tolist()] + sheet_df.values.tolist()
                }
                sheets_data.append(sheet_data)
                sheet_order.append(sheet_name)
            
            if not sheets_data:
                app.logger.error("No valid data found in any sheet")
                flash('No valid data found in Excel file', 'error')
                return redirect(url_for('index'))
                
            # Store the data in session or database
            # Add code here to persist the data
                
            flash(f'Successfully processed {len(sheets_data)} sheets', 'success')
            return redirect(url_for('index'))
            
    except Exception as e:
        app.logger.exception("Error processing upload")
        flash(f'Error processing file: {str(e)}', 'error')
        
    return redirect(url_for('index'))

@app.route('/update_header_color', methods=['POST'])
@login_required
def update_header_color():
    try:
        data = request.get_json()
        sheet_name = data.get('sheet_name')
        header_indices = data.get('headers', [])
        color = data.get('color')
        
        if not sheet_name or not header_indices or not color:
            return jsonify({'success': False, 'error': 'Missing required data'})
        
        sheet = Sheet.query.filter_by(user_id=current_user.id, sheet_name=sheet_name).first()
        if not sheet:
            return jsonify({'success': False, 'error': 'Sheet not found'})
        
        # Parse the sheet data
        sheet_data = json.loads(sheet.data) if isinstance(sheet.data, str) else sheet.data
        
        # Check if sheet_data is a dict with format_metadata
        if isinstance(sheet_data, dict):
            format_metadata = sheet_data.get('format_metadata', {})
        else:
            # Convert to the new format if it's just an array
            format_metadata = {}
            sheet_data = {
                'data': sheet_data,
                'format_metadata': format_metadata,
                'dimensions': {}
            }
        
        # Update color for each header
        for col_index in header_indices:
            cell_ref = f"{chr(65 + col_index)}1"  # A1, B1, etc.
            
            if cell_ref not in format_metadata:
                format_metadata[cell_ref] = {}
                
            format_metadata[cell_ref]['bg_color'] = color
        
        # Update the sheet data
        sheet_data['format_metadata'] = format_metadata
        sheet.data = json.dumps(sheet_data)
        
        # Update modified timestamp
        sheet.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({'success': True})
    
    except Exception as e:
        print(f"Error in update_header_color: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/merge_cells', methods=['POST'])
@login_required
def merge_cells():
    try:
        data = request.get_json()
        sheet_name = data.get('sheet_name')
        min_row = data.get('min_row')
        max_row = data.get('max_row')
        min_col = data.get('min_col')
        max_col = data.get('max_col')
        content = data.get('content', '')
        
        if not all([sheet_name is not None, min_row is not None, max_row is not None, 
                  min_col is not None, max_col is not None]):
            return jsonify({'success': False, 'error': 'Missing required data'})
        
        sheet = Sheet.query.filter_by(user_id=current_user.id, sheet_name=sheet_name).first()
        if not sheet:
            return jsonify({'success': False, 'error': 'Sheet not found'})
        
        # Parse the sheet data
        sheet_data = json.loads(sheet.data) if isinstance(sheet.data, str) else sheet.data
        
        # Check if sheet_data is a dict with merged_cells
        if isinstance(sheet_data, dict):
            merged_cells = sheet_data.get('merged_cells', [])
        else:
            # Convert to the new format if it's just an array
            merged_cells = []
            sheet_data = {
                'data': sheet_data,
                'format_metadata': {},
                'dimensions': {},
                'merged_cells': merged_cells
            }
        
        # Add new merged cell range
        merged_cells.append({
            'first_row': min_row,
            'last_row': max_row,
            'first_col': min_col,
            'last_col': max_col,
            'content': content
        })
        
        # Update the sheet data
        sheet_data['merged_cells'] = merged_cells
        sheet.data = json.dumps(sheet_data)
        
        # Update modified timestamp
        sheet.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({'success': True})
    
    except Exception as e:
        print(f"Error in merge_cells: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/remove_merge_cells', methods=['POST'])
@login_required
def remove_merge_cells():
    try:
        data = request.get_json()
        sheet_name = data.get('sheet_name')
        min_row = data.get('min_row')
        max_row = data.get('max_row')
        min_col = data.get('min_col')
        max_col = data.get('max_col')
        
        if not all([sheet_name is not None, min_row is not None, max_row is not None, 
                  min_col is not None, max_col is not None]):
            return jsonify({'success': False, 'error': 'Missing required data'})
        
        sheet = Sheet.query.filter_by(user_id=current_user.id, sheet_name=sheet_name).first()
        if not sheet:
            return jsonify({'success': False, 'error': 'Sheet not found'})
        
        # Parse the sheet data
        sheet_data = json.loads(sheet.data) if isinstance(sheet.data, str) else sheet.data
        
        # Check if sheet_data has merged cells
        if isinstance(sheet_data, dict) and 'merged_cells' in sheet_data:
            # Filter out the merge that matches these coordinates
            merged_cells = sheet_data['merged_cells']
            sheet_data['merged_cells'] = [
                merge for merge in merged_cells
                if not (merge['first_row'] == min_row and 
                       merge['last_row'] == max_row and
                       merge['first_col'] == min_col and
                       merge['last_col'] == max_col)
            ]
            
            # Save back to database
            sheet.data = json.dumps(sheet_data)
            sheet.updated_at = datetime.utcnow()
            db.session.commit()
            
            return jsonify({'success': True})
        
        return jsonify({'success': False, 'error': 'No merged cells found in sheet data'})
    
    except Exception as e:
        print(f"Error in remove_merge_cells: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/download_excel')
@login_required
def download_excel():
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import json
        import gc
        
        # Force garbage collection before processing
        gc.collect()
        
        # Get user sheets
        user_sheets = Sheet.query.filter_by(user_id=current_user.id).all()
        
        if not user_sheets:
            flash('No sheets found to download', 'warning')
            return redirect(url_for('index'))
        
        # Create a new workbook
        workbook = Workbook()
        
        # Remove default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        
        app.logger.info(f"Starting to process {len(user_sheets)} sheets")
        
        # Create a record for download history
        download_data = []
        
        # Process each sheet
        for idx, sheet_db in enumerate(user_sheets):
            try:
                # Log progress
                app.logger.info(f"Processing sheet {idx+1}/{len(user_sheets)}: {sheet_db.sheet_name}")
                
                # Ensure valid sheet name (Excel limits to 31 chars)
                safe_sheet_name = sheet_db.sheet_name[:31]
                
                # Create worksheet
                worksheet = workbook.create_sheet(title=safe_sheet_name)
                
                # Parse data from JSON
                if isinstance(sheet_db.data, str):
                    try:
                        sheet_data = json.loads(sheet_db.data)
                    except json.JSONDecodeError:
                        app.logger.error(f"JSON decode error for sheet: {sheet_db.sheet_name}")
                        flash(f'Error parsing data for sheet: {sheet_db.sheet_name}', 'error')
                        continue
                else:
                    sheet_data = sheet_db.data
                
                # Handle both old array format and new dictionary format
                if isinstance(sheet_data, list):
                    # Old format: direct array of rows
                    data = sheet_data
                    format_metadata = {}
                    dimensions = {}
                    merged_cells = []
                elif isinstance(sheet_data, dict):
                    # New format: object with data and metadata
                    data = sheet_data.get('data', [])
                    format_metadata = sheet_data.get('format_metadata', {})
                    dimensions = sheet_data.get('dimensions', {})
                    merged_cells = sheet_data.get('merged_cells', [])
                else:
                    app.logger.error(f"Unknown data format for sheet: {sheet_db.sheet_name}")
                    continue
                
                # Write data to worksheet
                if not data or not isinstance(data, list):
                    app.logger.warning(f"No valid data for sheet: {sheet_db.sheet_name}")
                    continue
                
                # Add to download record
                download_data.append({
                    'sheet_name': sheet_db.sheet_name,
                    'row_count': len(data),
                    'col_count': len(data[0]) if data and len(data) > 0 else 0
                })
                
                # Process in chunks to handle large sheets
                CHUNK_SIZE = 1000
                
                # Write header first (row 0)
                if data and len(data) > 0:
                    for col_idx, header in enumerate(data[0]):
                        cell = worksheet.cell(row=1, column=col_idx + 1)
                        cell.value = header
                        
                        # Apply header formatting 
                        cell.font = Font(bold=True, size=12)
                        cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Write data rows in chunks
                for start_idx in range(1, len(data), CHUNK_SIZE):
                    end_idx = min(start_idx + CHUNK_SIZE, len(data))
                    chunk = data[start_idx:end_idx]
                    
                    for row_idx, row_data in enumerate(chunk, start=start_idx + 1):
                        for col_idx, cell_value in enumerate(row_data):
                            # Handle Excel max columns (16384)
                            if col_idx >= 16384:
                                app.logger.warning(f"Exceeded Excel column limit in sheet {sheet_db.sheet_name}")
                                break
                                
                            cell = worksheet.cell(row=row_idx, column=col_idx + 1)
                            
                            # Parse formulas
                            if isinstance(cell_value, str) and cell_value.startswith('='):
                                cell.value = cell_value
                            else:
                                cell.value = cell_value
                
                # Set column widths
                column_widths = dimensions.get('column_widths', {})
                for col_idx_str, width in column_widths.items():
                    try:
                        col_idx = int(col_idx_str)
                        worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = width
                    except (ValueError, KeyError) as e:
                        app.logger.warning(f"Error setting column width: {e}")
                
                # Set row heights
                row_heights = dimensions.get('row_heights', {})
                for row_idx_str, height in row_heights.items():
                    try:
                        row_idx = int(row_idx_str)
                        worksheet.row_dimensions[row_idx + 1].height = height
                    except (ValueError, KeyError) as e:
                        app.logger.warning(f"Error setting row height: {e}")
                
                # Apply merged cells
                for merge in merged_cells:
                    try:
                        first_row = merge.get('first_row', 0)
                        last_row = merge.get('last_row', 0)
                        first_col = merge.get('first_col', 0)
                        last_col = merge.get('last_col', 0)
                        
                        # Excel uses 1-based indexing for cells
                        merge_range = f"{get_column_letter(first_col + 1)}{first_row + 1}:{get_column_letter(last_col + 1)}{last_row + 1}"
                        worksheet.merge_cells(merge_range)
                        
                        # Set content in the merged cell
                        if 'content' in merge:
                            worksheet.cell(row=first_row + 1, column=first_col + 1).value = merge.get('content', '')
                    except Exception as merge_error:
                        app.logger.warning(f"Error applying merge: {merge_error}")
                
            except Exception as sheet_error:
                app.logger.error(f"Error processing sheet {sheet_db.sheet_name}: {str(sheet_error)}")
                continue
        
        # Save workbook to bytes buffer
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Record the download in history
        try:
            history = DownloadHistory(
                user_id=current_user.id,
                filename=f"export_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
                sheet_data=json.dumps(download_data),
                downloaded_at=datetime.utcnow()
            )
            db.session.add(history)
            db.session.commit()
        except Exception as history_error:
            app.logger.warning(f"Could not save download history: {str(history_error)}")
        
        # Clean up to free memory
        del workbook
        gc.collect()
        
        # Send the file
        return send_file(
            output,
            as_attachment=True,
            download_name=f"Excel_Export_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        app.logger.error(f"Error generating Excel file: {str(e)}", exc_info=True)
        flash(f"Error generating Excel file: {str(e)}", "error")
        return redirect(url_for('index'))

@app.route('/update_sheet_format/<sheet_name>', methods=['POST'])
@login_required
def update_sheet_format(sheet_name):
    global sheets_data
    sheet = next((s for s in sheets_data if s['sheet_name'] == sheet_name), None)
    if not sheet:
        flash(f'Sheet "{sheet_name}" not found.', 'error')
        return redirect(url_for('index'))
    
    try:
        # Update formatting
        sheet['formatting'].update({
            'font': {
                'name': request.form.get('font_name', 'Arial'),
                'size': int(request.form.get('font_size', 11)),
                'bold': request.form.get('font_bold', 'off') == 'on',
                'italic': request.form.get('font_italic', 'off') == 'on',
                'color': request.form.get('font_color', '#000000')
            },
            'cell': {
                'alignment': request.form.get('cell_alignment', 'left'),
                'vertical_alignment': request.form.get('vertical_alignment', 'center'),
                'wrap_text': request.form.get('text_wrap', 'off') == 'on',
                'bg_color': request.form.get('cell_bg_color', '#FFFFFF')
            },
            'header': {
                'bg_color': request.form.get('header_bg_color', '#F0E68C'),
                'bold': request.form.get('header_bold', 'on') == 'on',
                'border': request.form.get('header_border', 'on') == 'on'
            },
            'border': {
                'style': request.form.get('border_style', 'thin'),
                'color': request.form.get('border_color', '#000000'),
                'all_borders': request.form.get('all_borders', 'off') == 'on',
                'outline': request.form.get('outline_only', 'off') == 'on'
            },
            'column_width': float(request.form.get('column_width', 15)),
            'row_height': float(request.form.get('row_height', 15))
        })
        
        flash(f'Formatting updated for sheet "{sheet_name}".', 'success')
    except Exception as e:
        app.logger.error(f"Error updating format for sheet {sheet_name}: {str(e)}")
        flash('Error updating sheet formatting.', 'error')
    return redirect(url_for('index'))

@app.route('/save_sheet_data', methods=['POST'])
@login_required
def save_sheet_data():
    try:
        data = request.get_json()
        sheet_name = data.get('sheet_name')
        sheet_data = data.get('data')
        format_info = data.get('format_info', {})
        
        sheet = Sheet.query.filter_by(user_id=current_user.id, sheet_name=sheet_name).first()
        if not sheet:
            return jsonify({'success': False, 'error': 'Sheet not found'})
        
        # Prepare the complete sheet data structure
        full_data = {
            'data': sheet_data,
            'format_info': format_info
        }
        
        # Save to database
        sheet.data = json.dumps(full_data)
        sheet.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        app.logger.error(f"Error saving sheet data: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/history')
@login_required
def view_history():
    downloads = DownloadHistory.query.filter_by(user_id=current_user.id)\
               .order_by(DownloadHistory.downloaded_at.desc()).all()
    return render_template('history.html', downloads=downloads)

@app.route('/add_to_templates/<int:download_id>', methods=['POST'])
@login_required
def add_to_templates(download_id):
    download = DownloadHistory.query.get_or_404(download_id)
    if download.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        # Parse sheet data from history
        sheet_data_json = json.loads(download.sheet_data) if isinstance(download.sheet_data, str) else download.sheet_data
        
        # Add each sheet to custom templates
        for sheet_item in sheet_data_json:
            sheet_data = json.loads(sheet_item['data']) if isinstance(sheet_item['data'], str) else sheet_item['data']
            
            # Create template from sheet
            template_id = f"template_{download.id}_{len(custom_templates)}"
            template_name = f"Template from {download.filename}"
            
            # Get headers and sample data
            headers = sheet_data[0] if sheet_data and len(sheet_data) > 0 else []
            sample_data = sheet_data[1:3] if sheet_data and len(sheet_data) > 1 else []
            
            # Add to templates
            custom_templates.append({
                'id': template_id,
                'name': template_name,
                'default_sheet_name': sheet_item['sheet_name'],
                'headers': headers,
                'sample_data': sample_data,
                'formatting': {
                    'header_bg_color': '#F0E68C',  # Default yellow
                    'title_font_size': 14,
                    'title_bold': True
                }
            })
        
        return jsonify({'success': True})
    except Exception as e:
        app.logger.error(f"Error adding template: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/download_historic/<int:download_id>')
@login_required
def download_historic(download_id):
    """Download a previously downloaded Excel file from history"""
    try:
        # Get the download history record
        download = DownloadHistory.query.get_or_404(download_id)
        
        # Check if the download belongs to the current user
        if download.user_id != current_user.id:
            flash('You do not have permission to access this file', 'error')
            return redirect(url_for('view_history'))
            
        # Parse the sheet data
        sheet_data_json = json.loads(download.sheet_data) if isinstance(download.sheet_data, str) else download.sheet_data
        
        # Create a new workbook
        workbook = Workbook()
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook.active)
            
        # Process each sheet from history
        for sheet_info in sheet_data_json:
            sheet_name = sheet_info.get('sheet_name', 'Sheet')
            
            # Fetch actual sheet data from database if possible
            sheet = Sheet.query.filter_by(user_id=current_user.id, sheet_name=sheet_name).first()
            
            if sheet:
                # We found the original sheet, use its current data
                sheet_data = json.loads(sheet.data) if isinstance(sheet.data, str) else sheet.data
                
                # Extract actual data array if needed
                if isinstance(sheet_data, dict) and 'data' in sheet_data:
                    data = sheet_data['data']
                    format_metadata = sheet_data.get('format_metadata', {})
                    merged_cells = sheet_data.get('merged_cells', [])
                else:
                    data = sheet_data
                    format_metadata = {}
                    merged_cells = []
            else:
                # Sheet doesn't exist anymore, create a simple placeholder
                ws = workbook.create_sheet(title=sheet_name[:31])  # Excel limits sheet names to 31 chars
                ws.append(['This sheet is no longer available'])
                continue
                
            # Create worksheet
            ws = workbook.create_sheet(title=sheet_name[:31])
            
            # Write data
            if data:
                for row_idx, row_data in enumerate(data):
                    for col_idx, cell_value in enumerate(row_data):
                        cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
                        
                        # Handle formulas
                        if isinstance(cell_value, str) and cell_value.startswith('='):
                            cell.value = cell_value
                        else:
                            cell.value = cell_value
                            
                        # Apply formatting if available
                        cell_ref = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                        if cell_ref in format_metadata:
                            fmt = format_metadata[cell_ref]
                            
                            # Apply background color
                            if 'bg_color' in fmt:
                                fill = PatternFill(start_color=fmt['bg_color'].replace('#', ''),
                                                  end_color=fmt['bg_color'].replace('#', ''), 
                                                  fill_type="solid")
                                cell.fill = fill
                                
                            # Apply text color
                            if 'text_color' in fmt:
                                font = Font(color=fmt['text_color'].replace('#', ''))
                                cell.font = font
                                
                            # Apply text alignment
                            if 'align' in fmt:
                                alignment = Alignment(horizontal=fmt['align'])
                                cell.alignment = alignment
                                
                            # Apply font formatting
                            if 'font_bold' in fmt and fmt['font_bold']:
                                if not cell.font:
                                    cell.font = Font(bold=True)
                                else:
                                    cell.font = Font(bold=True, color=cell.font.color)
            
            # Apply merges
            for merge in merged_cells:
                first_row = merge.get('first_row', 0)
                last_row = merge.get('last_row', 0)
                first_col = merge.get('first_col', 0)
                last_col = merge.get('last_col', 0)
                
                # Convert to Excel's 1-based indexing and column letters
                start_cell = f"{get_column_letter(first_col + 1)}{first_row + 1}"
                end_cell = f"{get_column_letter(last_col + 1)}{last_row + 1}"
                
                # Apply merge
                ws.merge_cells(f"{start_cell}:{end_cell}")
                
                # Set content in merged cell if available
                if 'content' in merge:
                    ws.cell(row=first_row + 1, column=first_col + 1).value = merge.get('content', '')
                    
        # Save to bytes buffer
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Send the file
        return send_file(
            output,
            as_attachment=True,
            download_name=download.filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        app.logger.error(f"Error downloading historic file: {str(e)}")
        flash(f"Error: {str(e)}", 'error')
        return redirect(url_for('view_history'))

@app.route('/profile')
@login_required
def profile():
    return render_template('profile.html', user=current_user)

@app.route('/edit_sheet/<sheet_name>')
@login_required
def edit_sheet(sheet_name):
    sheet = Sheet.query.filter_by(user_id=current_user.id, sheet_name=sheet_name).first()
    if not sheet:
        flash(f'Sheet "{sheet_name}" not found.', 'error')
        return redirect(url_for('index'))
    
    try:
        # Parse JSON data
        raw_data = json.loads(sheet.data) if isinstance(sheet.data, str) else sheet.data
        data = raw_data.get('data', []) if isinstance(raw_data, dict) else raw_data

        # Ensure valid data
        if not data:
            data = [[""]]

        return render_template('edit_sheet.html', sheet=sheet, data=data)
    except Exception as e:
        app.logger.error(f"Error editing sheet: {str(e)}")
        flash(f"Error loading sheet: {str(e)}", 'error')
        return redirect(url_for('index'))

@app.route('/sheet_preview/<sheet_name>')
@login_required
def sheet_preview(sheet_name):
    try:
        sheet = Sheet.query.filter_by(user_id=current_user.id, sheet_name=sheet_name).first()
        if not sheet:
            return jsonify({'success': False, 'error': 'Sheet not found'})
            
        # Parse JSON data
        sheet_data = json.loads(sheet.data) if isinstance(sheet.data, str) else sheet.data
        
        # Extract actual data array if needed
        if isinstance(sheet_data, dict) and 'data' in sheet_data:
            sheet_data = sheet_data['data']
            
        return jsonify({'success': True, 'data': sheet_data})
        
    except Exception as e:
        app.logger.error(f"Error loading sheet preview: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/update_sheet_dimensions', methods=['POST'])
@login_required
def update_sheet_dimensions():
    data = request.get_json()
    sheet_name = data.get('sheet_name')
    column_widths = data.get('column_widths', {})
    row_heights = data.get('row_heights', {})
    
    # Validate input
    if not sheet_name:
        return jsonify({'success': False, 'error': 'Sheet name not provided'})
    
    # Get the sheet
    sheet = Sheet.query.filter_by(user_id=current_user.id, sheet_name=sheet_name).first()
    if not sheet:
        return jsonify({'success': False, 'error': 'Sheet not found'})
    
    try:
        # Parse existing data
        sheet_data = json.loads(sheet.data) if isinstance(sheet.data, str) else sheet.data
        
        # Convert plain array to structured format if needed
        if isinstance(sheet_data, list):
            sheet_data = {'data': sheet_data}
        
        # Update or create dimensions property
        if 'dimensions' not in sheet_data:
            sheet_data['dimensions'] = {}
        
        if column_widths:
            sheet_data['dimensions']['column_widths'] = column_widths
        
        if row_heights:
            sheet_data['dimensions']['row_heights'] = row_heights
        
        # Save back to database
        sheet.data = json.dumps(sheet_data)
        db.session.commit()
        
        return jsonify({'success': True})
    
    except Exception as e:
        app.logger.error(f"Error updating dimensions: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/save_cell_format', methods=['POST'])
@login_required
def save_cell_format():
    try:
        data = request.get_json()
        sheet_name = data.get('sheet_name')
        cell_ref = data.get('cell_ref')
        format_data = data.get('format_data')
        
        if not sheet_name or not cell_ref or not format_data:
            return jsonify({'success': False, 'error': 'Missing required data'})
        
        sheet = Sheet.query.filter_by(user_id=current_user.id, sheet_name=sheet_name).first()
        if not sheet:
            return jsonify({'success': False, 'error': 'Sheet not found'})
        
        # Parse the sheet data
        sheet_data = json.loads(sheet.data) if isinstance(sheet.data, str) else sheet.data
        
        # Check if sheet_data is a dict with format_metadata
        if isinstance(sheet_data, dict):
            format_metadata = sheet_data.get('format_metadata', {})
        else:
            # Convert to the new format if it's just an array
            format_metadata = {}
            sheet_data = {
                'data': sheet_data,
                'format_metadata': format_metadata,
                'dimensions': {},
                'merged_cells': []
            }
        
        # Update or create format data for the cell
        if cell_ref not in format_metadata:
            format_metadata[cell_ref] = {}
        
        # Update format data
        for key, value in format_data.items():
            format_metadata[cell_ref][key] = value
        
        # Save back to sheet data
        sheet_data['format_metadata'] = format_metadata
        sheet.data = json.dumps(sheet_data)
        
        # Update modified timestamp
        sheet.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'success': True})
    
    except Exception as e:
        print(f"Error in save_cell_format: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

if __name__ == '__main__':
    # init_db()  # Only use this for initial setup or debugging!
    app.run(debug=True)